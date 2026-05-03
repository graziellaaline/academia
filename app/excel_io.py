# -*- coding: utf-8 -*-
"""
Importação e exportação de dados via Excel.
"""

import io
import logging
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database import get_conn

logger = logging.getLogger(__name__)

_PRETO   = "FF1a2840"
_VERMELHO = "FFe63946"
_BRANCO  = "FFFFFFFF"
_CINZA   = "FFF4F6FB"


def _cabecalho(ws, colunas, cor_fundo=_PRETO, cor_texto=_BRANCO):
    for col_idx, texto in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx, value=texto)
        cell.font      = Font(bold=True, color=cor_texto, size=10)
        cell.fill      = PatternFill("solid", fgColor=cor_fundo)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = Border(
            bottom=Side(style="thin", color="FFCCCCCC")
        )
    ws.row_dimensions[1].height = 20


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


# ── EXPORTAR ALUNOS ────────────────────────────────────────────────────────

def exportar_alunos() -> bytes:
    conn = get_conn()
    rows = conn.execute("""
        SELECT a.id, a.nome, a.cpf, a.data_nascimento, a.telefone, a.email,
               a.cep, a.logradouro, a.numero, a.bairro, a.cidade, a.uf,
               a.status, a.origem, a.observacoes, a.criado_em,
               (SELECT COUNT(*) FROM matriculas m
                WHERE m.aluno_id=a.id AND m.status='ativo') AS matriculas_ativas
        FROM alunos a
        ORDER BY a.nome
    """).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alunos"

    colunas = ["Nº Aluno", "Nome", "CPF", "Nascimento", "Telefone", "E-mail",
               "CEP", "Logradouro", "Número", "Bairro", "Cidade", "UF",
               "Status", "Origem", "Observações", "Cadastrado em", "Matrículas Ativas"]
    _cabecalho(ws, colunas)

    fill_par = PatternFill("solid", fgColor=_CINZA)
    for i, row in enumerate(rows, 2):
        d = dict(row)
        vals = list(d.values())
        vals[0] = f"#{d['id']:04d}"  # formata o ID como número de aluno
        for j, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.alignment = Alignment(vertical="center")
            if i % 2 == 0:
                cell.fill = fill_par

    _auto_width(ws)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── EXPORTAR PAGAMENTOS ────────────────────────────────────────────────────

def exportar_pagamentos(mes=None, ano=None) -> bytes:
    conn = get_conn()
    sql = """
        SELECT p.id, a.nome AS aluno, a.cpf, a.telefone,
               tp.nome AS plano, mod.nome AS modalidade,
               p.valor, p.data_vencimento, p.data_pagamento,
               p.forma, p.status, p.periodo_ref, p.observacoes
        FROM pagamentos p
        JOIN alunos a       ON a.id  = p.aluno_id
        JOIN matriculas m   ON m.id  = p.matricula_id
        JOIN tipos_plano tp ON tp.id = m.tipo_plano_id
        JOIN modalidades mod ON mod.id = m.modalidade_id
        WHERE 1=1
    """
    params = []
    if mes and ano:
        sql += " AND strftime('%m/%Y', p.data_vencimento)=?"
        params.append(f"{str(mes).zfill(2)}/{ano}")
    sql += " ORDER BY p.data_vencimento DESC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagamentos"

    colunas = ["ID", "Aluno", "CPF", "Telefone", "Plano", "Modalidade",
               "Valor (R$)", "Vencimento", "Data Pagamento", "Forma", "Status",
               "Período Ref.", "Observações"]
    _cabecalho(ws, colunas)

    fill_par  = PatternFill("solid", fgColor=_CINZA)
    fill_pago = PatternFill("solid", fgColor="FFd4edda")
    fill_venc = PatternFill("solid", fgColor="FFf8d7da")

    for i, row in enumerate(rows, 2):
        d = dict(row)
        vals = list(d.values())
        status = d.get("status", "")
        for j, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.alignment = Alignment(vertical="center")
        # colorir por status
        fill = (fill_pago if status == "pago"
                else fill_venc if status == "vencido"
                else fill_par if i % 2 == 0 else None)
        if fill:
            for j in range(1, len(vals) + 1):
                ws.cell(row=i, column=j).fill = fill

    _auto_width(ws)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── IMPORTAR ALUNOS ────────────────────────────────────────────────────────

def importar_alunos(conteudo_bytes: bytes) -> dict:
    """
    Lê planilha Excel e insere/atualiza alunos.
    Colunas obrigatórias: Nome, Telefone
    Colunas opcionais:    CPF, Nascimento, E-mail, CEP, Logradouro, Número,
                          Bairro, Cidade, UF, Observações
    Retorna dict com {'inseridos': N, 'erros': [...]}
    """
    from app.alunos import criar_aluno, _fmt_cpf, _fmt_nome_aluno

    wb = openpyxl.load_workbook(io.BytesIO(conteudo_bytes), data_only=True)
    ws = wb.active

    headers = [str(c.value or "").strip().lower() for c in ws[1]]
    _map = {
        "nome":          ["nome", "name", "aluno"],
        "cpf":           ["cpf"],
        "data_nascimento": ["nascimento", "data_nascimento", "data nascimento", "dt_nascimento"],
        "telefone":      ["telefone", "fone", "celular", "whatsapp"],
        "email":         ["email", "e-mail"],
        "cep":           ["cep"],
        "logradouro":    ["logradouro", "endereço", "endereco", "rua", "av"],
        "numero":        ["numero", "número", "num"],
        "bairro":        ["bairro"],
        "cidade":        ["cidade", "municipio", "município"],
        "uf":            ["uf", "estado"],
        "observacoes":   ["observações", "observacoes", "obs"],
    }

    def _col(campo):
        for alias in _map[campo]:
            if alias in headers:
                return headers.index(alias)
        return None

    idx = {campo: _col(campo) for campo in _map}

    if idx["nome"] is None:
        return {"inseridos": 0, "erros": ["Coluna 'Nome' não encontrada na planilha."]}

    inseridos = 0
    erros = []
    ids_inseridos = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        def _val(campo):
            i = idx[campo]
            if i is None or i >= len(row):
                return ""
            v = row[i]
            return str(v).strip() if v is not None else ""

        nome = _val("nome")
        if not nome:
            continue
        try:
            novo_id = criar_aluno({
                "nome":            _fmt_nome_aluno(nome),
                "cpf":             _fmt_cpf(_val("cpf")),
                "data_nascimento": _val("data_nascimento"),
                "telefone":        _val("telefone"),
                "email":           _val("email"),
                "cep":             _val("cep"),
                "logradouro":      _val("logradouro"),
                "numero":          _val("numero"),
                "bairro":          _val("bairro"),
                "cidade":          _val("cidade"),
                "uf":              _val("uf"),
                "endereco":        _val("logradouro"),
                "observacoes":     _val("observacoes"),
                "origem":          "importacao_excel",
            })
            ids_inseridos.append(novo_id)
            inseridos += 1
        except Exception as e:
            erros.append(f"Linha {row_num} ({nome}): {e}")

    primeiro = f"#{min(ids_inseridos):04d}" if ids_inseridos else "—"
    ultimo   = f"#{max(ids_inseridos):04d}" if ids_inseridos else "—"
    return {"inseridos": inseridos, "erros": erros,
            "primeiro": primeiro, "ultimo": ultimo}


# ── MODELO DE IMPORTAÇÃO ───────────────────────────────────────────────────

def gerar_modelo_importacao() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Importar Alunos"

    colunas = ["Nome", "CPF", "Nascimento", "Telefone", "E-mail",
               "CEP", "Logradouro", "Número", "Bairro", "Cidade", "UF", "Observações"]
    _cabecalho(ws, colunas)

    # Linha de exemplo
    ws.append(["João da Silva", "123.456.789-00", "1990-05-20",
               "(11) 99999-9999", "joao@email.com",
               "01310-100", "Av. Paulista", "1000", "Bela Vista", "São Paulo", "SP", ""])

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
